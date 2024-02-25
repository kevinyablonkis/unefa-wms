from openpyxl import *
import _CopyPasteData as cpd
import os

#Instanciación de las clases necesarias para los procesos de lectura y escritura de la información.
paste_data = cpd.PasteData
paste_style = cpd.CopyPasteStyles
read_json = cpd.JSONReader

#Ruta para los archivos json para extraer los datos.
root_dir = os.path.dirname(__file__)
route = os.path.join(root_dir, "json/data-products.json")
route_name = os.path.join(root_dir, "json/data-personal.json")

#Parámetros para extraer los datos en un nuevo Array de los productos y los nombres.
PARAMETER = ["Product_ID", "Code_Subgrupo", "Stock", "Serial", "Description", "AC", "DC"]
DATA_TO_INSERT = read_json(route).get_value_parameters(PARAMETER)
NAME = ["admin", "admbiene", "decano", "date"]
NAME_TO_INSERT = read_json(route_name).get_value_parameters_name(NAME)
RANGE_FORMAT = "A1:K42"

#Variables de control usadas a lo largo del programa.
rango_a = 1
rango_k = 42
rd1 = 10
rd2 = 34
rname = 40
rdate = 3
sr = 0
er = 0

#Función para pegar todos los datos en la hoja dependiendo de su sub-grupo donde los rangos son dinámicos y van aumentando su valor para agregar correctamente los datos.
def data_paste(rd1, rd2, rname, rdate, er, sr, rango_a, rango_k):
    for i in new_list[1]:
        ranges = [f"A{rd1}:A{rd2}", f"B{rd1}:B{rd2}", f"C{rd1}:C{rd2}", f"D{rd1}:D{rd2}", f"E{rd1}:E{rd2}", f"J{rd1}:J{rd2}", f"K{rd1}:K{rd2}"]

        if er < len(new_list[1]):
            ws_active[f"C{rname}"].value = NAME_TO_INSERT[0]
            ws_active[f"F{rname}"].value = NAME_TO_INSERT[1]
            ws_active[f"I{rname}"].value = NAME_TO_INSERT[2]
            ws_active[f"A{rdate}"].value = f"FECHA:{NAME_TO_INSERT[3]}"
            er += 25
            er = min(er, len(new_list[1]))
            data = [[subdata[j] for j in range(sr, er)] for subdata in new_list]
            paste_data(ws_active).add_all_data(ranges, data)

            sr += 25
            rd1 += 42
            rd2 += 42
            rname += 42
            rdate += 42

            if er < len(new_list[1]):
                rango_a += 42
                rango_k += 42
                paste_style(ws_format, ws_active).copy_paste_all(RANGE_FORMAT, f"A{rango_a}:K84{rango_k}")

#Programa.
if __name__ == "__main__":
    #Cargar el cuaderno y la hoja donde se encuentra el formato a usar.
    wb_format = load_workbook(os.path.join(root_dir, "sheets/formato.xlsx"))
    ws_format = wb_format["Hoja1"]

    #Creación de un cuaderno piloto para evitar pérdidas.
    wb_new = Workbook()
    wb_new.save(os.path.join(root_dir, "sheets/_tempSheet.xlsx"))

    #Cargar el cuaderno donde se pegarán los datos y estilos.
    wb_rendition = load_workbook(os.path.join(root_dir, "sheets/_tempSheet.xlsx"))
    ws_active = wb_rendition.active

    #Diccionario para agrupar los productos por sub-grupos.
    data_dict = {}
    for i in range(len(DATA_TO_INSERT[1])):
        key = DATA_TO_INSERT[1][i]
        if key not in data_dict:
            data_dict[key] = []
        data_dict[key].append([row[i] for row in DATA_TO_INSERT])

    #Bucle para el pegado de los datos por subgrupos, donde la variable k toma como valor el número del subgrupo actual en data_dict, luego se ordenan los productos en new_list para facilitar su agregado y se agregan todos los productos en la hoja actual.
    for k in data_dict:
        new_list = list(zip(*data_dict[k]))

        for i in range(len(new_list[0])):
            new_list[0] = list(new_list[0])
            new_list[0][i] = 2
        new_list[0] = tuple(new_list[0])

        ws_active = wb_rendition.create_sheet(title=f"SUB-GRUPO 0{k}")
        paste_style(ws_format, ws_active).copy_paste_all(RANGE_FORMAT, f"A{rango_a}:K84{rango_k}")
        data_paste(rd1, rd2, rname, rdate, er, sr, rango_a, rango_k)

    #Guardar el cuaderno.
    wb_rendition.save(os.path.join(root_dir, f"sheets/Rendición General Anual.xlsx"))
